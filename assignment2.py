"""
Assignment 2: Translation with LLM
Model Used : Claude claude-sonnet-4-6 (Anthropic API)
Task       : Translate 100 English sentences to Hindi
Metrics    : BLEU, CHRF, TER (via sacrebleu)

Requirements:
    pip install anthropic sacrebleu pandas openpyxl

Setup:
    Set your Anthropic API key as an environment variable:
        export ANTHROPIC_API_KEY="your_key_here"
    Or replace the os.environ.get(...) call below with your key directly.
"""

import os
import re
import time
import json
import pandas as pd
import sacrebleu
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
API_KEY         = os.environ.get("ANTHROPIC_API_KEY", "YOUR_API_KEY_HERE")
INPUT_EXCEL     = "assignment1_output.xlsx"   # Output from Assignment 1
OUTPUT_EXCEL    = "assignment2_translations.xlsx"
OUTPUT_METRICS  = "assignment2_metrics.txt"
NUM_SENTENCES   = 100
BATCH_SIZE      = 10
RANDOM_SEED     = 42


# ─────────────────────────────────────────────
# STEP 1: Load 100 sentences from Assignment 1
# ─────────────────────────────────────────────
def load_sentences(path, n=100, seed=42):
    df = pd.read_excel(path)
    df.columns = [
        "English Sentences", "Hindi Sentences",
        "Word Count (English)", "Word Count (Hindi)", "Difference"
    ]
    sample = df.sample(n=n, random_state=seed).reset_index(drop=True)
    print(f"[INFO] Loaded {n} sentences from {path}")
    return sample["English Sentences"].tolist(), sample["Hindi Sentences"].tolist()


# ─────────────────────────────────────────────
# STEP 2: Translate using Claude claude-sonnet-4-6
# ─────────────────────────────────────────────
def translate_with_claude(sentences, api_key, batch_size=10):
    client = anthropic.Anthropic(api_key=api_key)
    translations = []

    for i in range(0, len(sentences), batch_size):
        batch = sentences[i:i + batch_size]
        numbered = "\n".join([f"{j+1}. {s}" for j, s in enumerate(batch)])

        prompt = (
            "Translate the following English sentences into Hindi. "
            "Return ONLY a JSON array of translated strings in the same order, "
            "with no extra text, explanation, or markdown formatting.\n\n"
            f"{numbered}"
        )

        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )

        raw = message.content[0].text.strip()
        # Strip markdown fences if present
        raw = re.sub(r"^```json|^```|```$", "", raw, flags=re.MULTILINE).strip()
        batch_translations = json.loads(raw)
        translations.extend(batch_translations)

        print(f"  Translated {min(i + batch_size, len(sentences))}/{len(sentences)} sentences...")
        time.sleep(0.3)  # Rate limit buffer

    return translations


# ─────────────────────────────────────────────
# STEP 3: Compute BLEU, CHRF, TER
# ─────────────────────────────────────────────
def compute_metrics(predictions, references):
    refs = [references]
    bleu = sacrebleu.corpus_bleu(predictions, refs).score
    chrf = sacrebleu.corpus_chrf(predictions, refs).score
    ter  = sacrebleu.corpus_ter(predictions, refs).score
    return bleu, chrf, ter


# ─────────────────────────────────────────────
# STEP 4: Save metrics to .txt
# ─────────────────────────────────────────────
def save_metrics(bleu, chrf, ter, path):
    with open(path, "w", encoding="utf-8") as f:
        f.write("=" * 50 + "\n")
        f.write("Assignment 2: Translation Evaluation Metrics\n")
        f.write("=" * 50 + "\n\n")
        f.write(f"Model Used  : Claude claude-sonnet-4-6 (Anthropic)\n")
        f.write(f"Sentences   : 100 (sampled from Assignment 1 output)\n\n")
        f.write(f"BLEU Score  : {bleu:.4f}\n")
        f.write(f"CHRF Score  : {chrf:.4f}\n")
        f.write(f"TER Score   : {ter:.4f}\n\n")
        f.write("Metric Notes:\n")
        f.write("  BLEU  : Higher is better (0-100). Measures n-gram overlap precision.\n")
        f.write("  CHRF  : Higher is better (0-100). Character n-gram F-score.\n")
        f.write("  TER   : Lower is better. Minimum edits needed to match reference.\n")
    print(f"[INFO] Metrics saved: {path}")


# ─────────────────────────────────────────────
# STEP 5: Save translations to Excel
# ─────────────────────────────────────────────
def clean_text(text):
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', str(text))

def save_translations_excel(english_sentences, predicted_hindi, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"

    headers = ["Original English Sentence", "Model-Generated Hindi Translation"]
    header_font  = Font(bold=True, name="Arial", color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, (eng, hi) in enumerate(zip(english_sentences, predicted_hindi), start=2):
        ws.cell(row=row_idx, column=1, value=clean_text(eng)).font = Font(name="Arial", size=10)
        ws.cell(row=row_idx, column=2, value=clean_text(hi)).font  = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 70
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:B1"

    wb.save(path)
    print(f"[INFO] Translations saved: {path}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print("=== Assignment 2: Translation with LLM ===\n")

    # Step 1: Load sentences
    english_sentences, reference_hindi = load_sentences(INPUT_EXCEL, n=NUM_SENTENCES, seed=RANDOM_SEED)

    # Step 2: Translate
    print("[INFO] Translating with Claude claude-sonnet-4-6...")
    predicted_hindi = translate_with_claude(english_sentences, API_KEY, batch_size=BATCH_SIZE)
    print(f"[INFO] Translation complete. Got {len(predicted_hindi)} translations.\n")

    # Step 3: Metrics
    print("[INFO] Computing evaluation metrics...")
    bleu, chrf, ter = compute_metrics(predicted_hindi, reference_hindi)
    print(f"  BLEU  : {bleu:.4f}")
    print(f"  CHRF  : {chrf:.4f}")
    print(f"  TER   : {ter:.4f}\n")

    # Step 4: Save metrics
    save_metrics(bleu, chrf, ter, OUTPUT_METRICS)

    # Step 5: Save Excel
    save_translations_excel(english_sentences, predicted_hindi, OUTPUT_EXCEL)

    print("\n✅ Assignment 2 Complete!")
    print(f"   Excel  : {OUTPUT_EXCEL}")
    print(f"   Metrics: {OUTPUT_METRICS}")
