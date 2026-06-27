"""
Assignment 1: English-Hindi Dataset Processing and Analysis
Author      : Sayan H. Mandal
Dataset     : Dataset_English_Hindi.csv (Kaggle English-Hindi Parallel Corpus)

Steps:
    1. Load the CSV dataset
    2. Extract into two columns: English and Hindi
    3. Compute word counts for both languages
    4. Filter: word count in range [5, 50] for both languages
    5. Filter: word count difference (English - Hindi) in [-10, +10]
    6. Save cleaned dataset to Excel with 5 required columns

Requirements:
    pip install pandas openpyxl
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
INPUT_CSV   = "Dataset_English_Hindi.csv"
OUTPUT_FILE = "assignment1_output.xlsx"
MIN_WC      = 5
MAX_WC      = 50
MIN_DIFF    = -10
MAX_DIFF    = 10
MIN_ROWS    = 10000


# ─────────────────────────────────────────────
# STEP 1 & 2: Load Dataset & Extract Columns
# ─────────────────────────────────────────────
def load_dataset(path):
    df = pd.read_csv(path)
    df.columns = ["English Sentences", "Hindi Sentences"]
    df = df.dropna()
    df = df[
        (df["English Sentences"].str.strip() != "") &
        (df["Hindi Sentences"].str.strip() != "")
    ].reset_index(drop=True)
    print(f"[INFO] Loaded {len(df)} rows from '{path}'")
    return df


# ─────────────────────────────────────────────
# STEP 3: Word Count Analysis
# ─────────────────────────────────────────────
def compute_word_counts(df):
    df["Word Count (English)"] = df["English Sentences"].apply(lambda x: len(str(x).split()))
    df["Word Count (Hindi)"]   = df["Hindi Sentences"].apply(lambda x: len(str(x).split()))
    return df


# ─────────────────────────────────────────────
# STEP 4: Filter by Word Count Range [5, 50]
# ─────────────────────────────────────────────
def filter_by_word_count(df, min_wc=5, max_wc=50):
    filtered = df[
        (df["Word Count (English)"] >= min_wc) & (df["Word Count (English)"] <= max_wc) &
        (df["Word Count (Hindi)"]   >= min_wc) & (df["Word Count (Hindi)"]   <= max_wc)
    ].reset_index(drop=True)
    print(f"[INFO] After word count filter [{min_wc}, {max_wc}]: {len(filtered)} rows")
    return filtered


# ─────────────────────────────────────────────
# STEP 5: Word Count Difference Filter [-10, +10]
# ─────────────────────────────────────────────
def filter_by_difference(df, min_diff=-10, max_diff=10):
    df["Difference between Word Count (English) and Word Count (Hindi)"] = (
        df["Word Count (English)"] - df["Word Count (Hindi)"]
    )
    filtered = df[
        (df["Difference between Word Count (English) and Word Count (Hindi)"] >= min_diff) &
        (df["Difference between Word Count (English) and Word Count (Hindi)"] <= max_diff)
    ].reset_index(drop=True)
    print(f"[INFO] After difference filter [{min_diff}, {max_diff}]: {len(filtered)} rows")
    return filtered


# ─────────────────────────────────────────────
# STEP 6: Save to Excel
# ─────────────────────────────────────────────
def clean_text(text):
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', str(text))

def save_to_excel(df, path):
    columns = [
        "English Sentences",
        "Hindi Sentences",
        "Word Count (English)",
        "Word Count (Hindi)",
        "Difference between Word Count (English) and Word Count (Hindi)"
    ]
    df["English Sentences"] = df["English Sentences"].apply(clean_text)
    df["Hindi Sentences"]   = df["Hindi Sentences"].apply(clean_text)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Dataset"

    header_font  = Font(bold=True, name="Arial", color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align

    for row_idx, row in enumerate(df[columns].itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")

    for i, width in enumerate([60, 60, 22, 20, 50], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"[INFO] Output saved to: '{path}'")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print("=== Assignment 1: English-Hindi Dataset Processing ===\n")
    df = load_dataset(INPUT_CSV)
    df = compute_word_counts(df)
    df = filter_by_word_count(df, MIN_WC, MAX_WC)
    df = filter_by_difference(df, MIN_DIFF, MAX_DIFF)

    if len(df) < MIN_ROWS:
        print(f"[WARNING] Only {len(df)} rows — below {MIN_ROWS} minimum!")
    else:
        print(f"[INFO] ✅ {len(df)} rows — meets 10,000 row requirement")

    save_to_excel(df, OUTPUT_FILE)
    print(f"\n✅ Assignment 1 Complete! Output: {OUTPUT_FILE}")
